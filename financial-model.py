from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ──
BLUE = Font(color="0000FF", size=11, name="Arial")  # inputs
BLUE_BOLD = Font(color="0000FF", size=11, name="Arial", bold=True)
BLACK = Font(color="000000", size=11, name="Arial")  # formulas
BLACK_BOLD = Font(color="000000", size=11, name="Arial", bold=True)
HEADER = Font(color="FFFFFF", size=11, name="Arial", bold=True)
SECTION = Font(color="000000", size=12, name="Arial", bold=True)
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
DARK_FILL = PatternFill("solid", fgColor="1B2A3D")
LIGHT_FILL = PatternFill("solid", fgColor="F2F7FB")
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
RED_FILL = PatternFill("solid", fgColor="FFEBEE")
BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

EUR = '€#,##0.00;(€#,##0.00);"-"'
EUR_INT = '€#,##0;(€#,##0);"-"'
PCT = '0.0%'
NUM = '#,##0'
NUM_DEC = '#,##0.00'

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER
        cell.fill = DARK_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def style_input(cell, fmt=None):
    cell.font = BLUE
    if fmt: cell.number_format = fmt

def style_formula(cell, fmt=None):
    cell.font = BLACK
    if fmt: cell.number_format = fmt

def style_section(cell):
    cell.font = SECTION

# ════════════════════════════════════════
# SHEET 1: API Cost Per Scan
# ════════════════════════════════════════
ws1 = wb.active
ws1.title = "API Costs"
ws1.sheet_properties.tabColor = "1B2A3D"

ws1.column_dimensions["A"].width = 35
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 22

# Title
ws1["A1"] = "CMO.ie — API Cost Model"
ws1["A1"].font = Font(size=14, bold=True, name="Arial")
ws1["A2"] = "What does each scan actually cost us?"
ws1["A2"].font = Font(size=10, color="666666", name="Arial")

# Section: Token Pricing
r = 4
ws1.cell(r, 1, "MODEL TOKEN PRICING"); style_section(ws1.cell(r, 1))
r = 5
headers = ["Model / Provider", "Input (per 1M tokens)", "Output (per 1M tokens)", "Avg tokens/call", "Source"]
for i, h in enumerate(headers, 1):
    ws1.cell(r, i, h)
style_header_row(ws1, r, 5)

# Row data: model, input$/M, output$/M, avg_tokens, source
models = [
    ("Claude Sonnet 4 (Anthropic)", 3.00, 15.00, 1200, "anthropic.com/pricing, Apr 2026"),
    ("GPT-4o (OpenAI)", 2.50, 10.00, 1100, "openai.com/pricing, Apr 2026"),
    ("Gemini 2.0 Flash (Google)", 0.10, 0.40, 1000, "ai.google.dev/pricing, Apr 2026"),
    ("Perplexity Sonar Pro", 3.00, 15.00, 1200, "docs.perplexity.ai/pricing, Apr 2026"),
    ("Claude (analysis step)", 3.00, 15.00, 800, "Same key — used for brand detection"),
]

for i, (name, inp, out, tok, src) in enumerate(models):
    row = r + 1 + i
    ws1.cell(row, 1, name).font = BLACK
    style_input(ws1.cell(row, 2, inp), EUR)
    style_input(ws1.cell(row, 3, out), EUR)
    style_input(ws1.cell(row, 4, tok), NUM)
    ws1.cell(row, 5, src).font = Font(size=9, color="999999", name="Arial")

# Section: Cost Per Prompt×Model Check
r = 12
ws1.cell(r, 1, "COST PER PROMPT × MODEL CHECK"); style_section(ws1.cell(r, 1))
r = 13
for i, h in enumerate(["Step", "Tokens (est)", "Cost per call", "Notes"], 1):
    ws1.cell(r, i, h)
style_header_row(ws1, r, 4)

# Step 1: Query the model (simulate response)
ws1.cell(14, 1, "1. Query model (simulate response)").font = BLACK
ws1.cell(14, 2, "~1,200 tokens out").font = BLACK
# Cost = (input_tokens * input_price + output_tokens * output_price) / 1M
# Avg: ~200 input tokens, ~1000 output tokens
# Using Claude Sonnet: (200 * 3 + 1000 * 15) / 1,000,000
ws1.cell(14, 3).font = BLACK
ws1.cell(14, 3, "=ROUND((200*B6/1000000)+(1000*C6/1000000),5)")
ws1.cell(14, 3).number_format = '€#,##0.00000'
ws1.cell(14, 4, "~200 input + ~1000 output tokens (Claude simulating)").font = Font(size=9, color="666666", name="Arial")

# Step 2: Analyse response
ws1.cell(15, 1, "2. Analyse response (brand detection)").font = BLACK
ws1.cell(15, 2, "~800 tokens total").font = BLACK
ws1.cell(15, 3, "=ROUND((600*B10/1000000)+(200*C10/1000000),5)")
ws1.cell(15, 3).number_format = '€#,##0.00000'
ws1.cell(15, 4, "~600 input + ~200 output tokens (analysis)").font = Font(size=9, color="666666", name="Arial")

# Total per check
ws1.cell(16, 1, "TOTAL per prompt × model check").font = BLACK_BOLD
ws1.cell(16, 3, "=C14+C15")
ws1.cell(16, 3).font = BLACK_BOLD
ws1.cell(16, 3).number_format = '€#,##0.00000'

# ════════════════════════════════════════
# Section: Cost Per Daily Scan by Plan
# ════════════════════════════════════════
r = 18
ws1.cell(r, 1, "COST PER DAILY SCAN BY PLAN"); style_section(ws1.cell(r, 1))
r = 19
for i, h in enumerate(["Plan", "Prompts", "Models", "Checks per scan", "Cost per scan"], 1):
    ws1.cell(r, i, h)
style_header_row(ws1, r, 5)

plans = [
    ("Trial", 10, 3),
    ("Starter (€29/mo)", 25, 3),
    ("Pro (€79/mo)", 50, 5),
    ("Advanced (€199/mo)", 100, 5),
]

for i, (name, prompts, models_count) in enumerate(plans):
    row = r + 1 + i
    ws1.cell(row, 1, name).font = BLACK
    style_input(ws1.cell(row, 2, prompts), NUM)
    style_input(ws1.cell(row, 3, models_count), NUM)
    ws1.cell(row, 4, f"=B{row}*C{row}").font = BLACK
    ws1.cell(row, 4).number_format = NUM
    ws1.cell(row, 5, f"=D{row}*$C$16").font = BLACK
    ws1.cell(row, 5).number_format = EUR

# Monthly cost (30 days)
r = 25
ws1.cell(r, 1, "MONTHLY COST (30 daily scans)"); style_section(ws1.cell(r, 1))
r = 26
for i, h in enumerate(["Plan", "Daily cost", "Monthly cost (30d)", "Monthly revenue", "Gross margin"], 1):
    ws1.cell(r, i, h)
style_header_row(ws1, r, 5)

revenues = [0, 29, 79, 199]
for i in range(4):
    row = r + 1 + i
    src_row = 20 + i
    ws1.cell(row, 1, plans[i][0]).font = BLACK
    ws1.cell(row, 2, f"=E{src_row}").font = BLACK
    ws1.cell(row, 2).number_format = EUR
    ws1.cell(row, 3, f"=B{row}*30").font = BLACK
    ws1.cell(row, 3).number_format = EUR
    style_input(ws1.cell(row, 4, revenues[i]), EUR_INT)
    if revenues[i] > 0:
        ws1.cell(row, 5, f"=(D{row}-C{row})/D{row}").font = BLACK
    else:
        ws1.cell(row, 5, "N/A — trial").font = Font(size=10, color="999999", name="Arial")
    ws1.cell(row, 5).number_format = PCT

# Highlight margins
for i in range(4):
    row = r + 1 + i
    if revenues[i] > 0:
        ws1.cell(row, 5).fill = GREEN_FILL

# ════════════════════════════════════════
# SHEET 2: Revenue Projections
# ════════════════════════════════════════
ws2 = wb.create_sheet("Revenue Model")
ws2.sheet_properties.tabColor = "10B981"

ws2.column_dimensions["A"].width = 30
for c in range(2, 14):
    ws2.column_dimensions[get_column_letter(c)].width = 14

ws2["A1"] = "CMO.ie — 12 Month Revenue Projection"
ws2["A1"].font = Font(size=14, bold=True, name="Arial")
ws2["A2"] = "Flat monthly subscription model — all API costs included"
ws2["A2"].font = Font(size=10, color="666666", name="Arial")

# Month headers
r = 4
ws2.cell(r, 1, "ASSUMPTIONS")
style_section(ws2.cell(r, 1))
for m in range(1, 13):
    ws2.cell(r, m + 1, f"Month {m}").font = HEADER
    ws2.cell(r, m + 1).fill = DARK_FILL
    ws2.cell(r, m + 1).alignment = Alignment(horizontal="center")
ws2.cell(r, 1).fill = DARK_FILL
ws2.cell(r, 1).font = HEADER

# Customer growth assumptions
r = 5
ws2.cell(r, 1, "New trial signups / month").font = BLACK
for m in range(1, 13):
    # Start with 20, grow 15% monthly
    if m == 1:
        style_input(ws2.cell(r, m + 1, 20), NUM)
    else:
        col = get_column_letter(m + 1)
        prev = get_column_letter(m)
        ws2.cell(r, m + 1, f"=ROUND({prev}{r}*1.15,0)").font = BLACK
        ws2.cell(r, m + 1).number_format = NUM

r = 6
ws2.cell(r, 1, "Trial → Starter conversion %").font = BLACK
for m in range(1, 13):
    style_input(ws2.cell(r, m + 1, 0.12), PCT)
    ws2.cell(r, m + 1).fill = YELLOW_FILL

r = 7
ws2.cell(r, 1, "Starter → Pro upgrade % (monthly)").font = BLACK
for m in range(1, 13):
    style_input(ws2.cell(r, m + 1, 0.08), PCT)
    ws2.cell(r, m + 1).fill = YELLOW_FILL

r = 8
ws2.cell(r, 1, "Pro → Advanced upgrade % (monthly)").font = BLACK
for m in range(1, 13):
    style_input(ws2.cell(r, m + 1, 0.03), PCT)
    ws2.cell(r, m + 1).fill = YELLOW_FILL

r = 9
ws2.cell(r, 1, "Monthly churn % (all paid)").font = BLACK
for m in range(1, 13):
    style_input(ws2.cell(r, m + 1, 0.05), PCT)
    ws2.cell(r, m + 1).fill = YELLOW_FILL

# Active customers by plan
r = 11
ws2.cell(r, 1, "ACTIVE CUSTOMERS"); style_section(ws2.cell(r, 1))
for m in range(1, 13):
    ws2.cell(r, m + 1).fill = DARK_FILL
style_header_row(ws2, r, 13)

r = 12
ws2.cell(r, 1, "Trial users (active)").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    if m == 1:
        ws2.cell(r, m + 1, f"={col}5").font = BLACK
    else:
        prev = get_column_letter(m)
        # New signups + leftover trials (assume trials last 14 days, so ~50% still active)
        ws2.cell(r, m + 1, f"=ROUND({col}5 + {prev}{r}*0.5*(1-{col}6), 0)").font = BLACK
    ws2.cell(r, m + 1).number_format = NUM

r = 13
ws2.cell(r, 1, "Starter customers").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    if m == 1:
        ws2.cell(r, m + 1, f"=ROUND({col}12*{col}6, 0)").font = BLACK
    else:
        prev = get_column_letter(m)
        # Previous starters + new conversions - upgrades to pro - churn
        ws2.cell(r, m + 1, f"=MAX(0, ROUND({prev}{r}*(1-{col}9-{col}7) + {col}12*{col}6, 0))").font = BLACK
    ws2.cell(r, m + 1).number_format = NUM

r = 14
ws2.cell(r, 1, "Pro customers").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    if m == 1:
        ws2.cell(r, m + 1, f"=ROUND({col}13*{col}7, 0)").font = BLACK
    else:
        prev = get_column_letter(m)
        ws2.cell(r, m + 1, f"=MAX(0, ROUND({prev}{r}*(1-{col}9-{col}8) + {col}13*{col}7, 0))").font = BLACK
    ws2.cell(r, m + 1).number_format = NUM

r = 15
ws2.cell(r, 1, "Advanced customers").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    if m == 1:
        ws2.cell(r, m + 1, f"=ROUND({col}14*{col}8, 0)").font = BLACK
    else:
        prev = get_column_letter(m)
        ws2.cell(r, m + 1, f"=MAX(0, ROUND({prev}{r}*(1-{col}9) + {col}14*{col}8, 0))").font = BLACK
    ws2.cell(r, m + 1).number_format = NUM

r = 16
ws2.cell(r, 1, "Total paying customers").font = BLACK_BOLD
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}13+{col}14+{col}15").font = BLACK_BOLD
    ws2.cell(r, m + 1).number_format = NUM

# Revenue
r = 18
ws2.cell(r, 1, "MONTHLY REVENUE"); style_section(ws2.cell(r, 1))
style_header_row(ws2, r, 13)

r = 19
ws2.cell(r, 1, "Starter revenue (€29/mo)").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}13*29").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 20
ws2.cell(r, 1, "Pro revenue (€79/mo)").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}14*79").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 21
ws2.cell(r, 1, "Advanced revenue (€199/mo)").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}15*199").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 22
ws2.cell(r, 1, "TOTAL REVENUE").font = BLACK_BOLD
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}19+{col}20+{col}21").font = BLACK_BOLD
    ws2.cell(r, m + 1).number_format = EUR_INT
    ws2.cell(r, m + 1).fill = GREEN_FILL

# API Costs
r = 24
ws2.cell(r, 1, "MONTHLY API COSTS"); style_section(ws2.cell(r, 1))
style_header_row(ws2, r, 13)

# Reference cost per scan from Sheet 1
# Trial: 10 prompts × 3 models = 30 checks
# Starter: 25 × 3 = 75 checks
# Pro: 50 × 5 = 250 checks
# Advanced: 100 × 5 = 500 checks

r = 25
ws2.cell(r, 1, "Trial API cost (free — capped)").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    # Trial: 30 checks/day × cost_per_check × 14 days (trial period)
    ws2.cell(r, m + 1, f"={col}12 * 30 * 'API Costs'!$C$16 * 14").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 26
ws2.cell(r, 1, "Starter API cost").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}13 * 75 * 'API Costs'!$C$16 * 30").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 27
ws2.cell(r, 1, "Pro API cost").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}14 * 250 * 'API Costs'!$C$16 * 30").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 28
ws2.cell(r, 1, "Advanced API cost").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}15 * 500 * 'API Costs'!$C$16 * 30").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 29
ws2.cell(r, 1, "TOTAL API COST").font = BLACK_BOLD
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}25+{col}26+{col}27+{col}28").font = BLACK_BOLD
    ws2.cell(r, m + 1).number_format = EUR_INT
    ws2.cell(r, m + 1).fill = RED_FILL

# Gross profit
r = 31
ws2.cell(r, 1, "P&L SUMMARY"); style_section(ws2.cell(r, 1))
style_header_row(ws2, r, 13)

r = 32
ws2.cell(r, 1, "Total Revenue").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}22").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 33
ws2.cell(r, 1, "Total API Cost").font = BLACK
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}29").font = BLACK
    ws2.cell(r, m + 1).number_format = EUR_INT

r = 34
ws2.cell(r, 1, "GROSS PROFIT").font = BLACK_BOLD
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"={col}32-{col}33").font = BLACK_BOLD
    ws2.cell(r, m + 1).number_format = EUR_INT
    ws2.cell(r, m + 1).fill = GREEN_FILL

r = 35
ws2.cell(r, 1, "Gross Margin %").font = BLACK_BOLD
for m in range(1, 13):
    col = get_column_letter(m + 1)
    ws2.cell(r, m + 1, f"=IF({col}32>0,{col}34/{col}32,0)").font = BLACK_BOLD
    ws2.cell(r, m + 1).number_format = PCT

r = 37
ws2.cell(r, 1, "CUMULATIVE").font = SECTION
for m in range(1, 13):
    col = get_column_letter(m + 1)
    if m == 1:
        ws2.cell(37, m + 1, f"={col}34").font = BLACK
    else:
        prev = get_column_letter(m)
        ws2.cell(37, m + 1, f"={prev}37+{col}34").font = BLACK
    ws2.cell(37, m + 1).number_format = EUR_INT

# ════════════════════════════════════════
# SHEET 3: Plan Comparison
# ════════════════════════════════════════
ws3 = wb.create_sheet("Plan Comparison")
ws3.sheet_properties.tabColor = "F59E0B"

ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 16
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 16
ws3.column_dimensions["E"].width = 16

ws3["A1"] = "CMO.ie — Plan Comparison & Pricing Rationale"
ws3["A1"].font = Font(size=14, bold=True, name="Arial")

r = 3
for i, h in enumerate(["Feature", "Trial (Free)", "Starter (€29)", "Pro (€79)", "Advanced (€199)"], 1):
    ws3.cell(r, i, h)
style_header_row(ws3, r, 5)

features = [
    ("Monthly price", "€0", "€29", "€79", "€199"),
    ("Projects", "1", "1", "3", "Unlimited"),
    ("Prompts per project", "10", "25", "50", "100"),
    ("AI Models", "3", "3", "5", "5"),
    ("Daily scans", "14 days only", "Daily", "Daily", "Daily"),
    ("Checks per scan", "30", "75", "250", "500"),
    ("Action plan tier", "Gap analysis", "Gap analysis", "Strategy + actions", "Full + briefs"),
    ("Brief credits / month", "0", "5", "20", "Unlimited"),
    ("Team members", "1", "1", "3", "Unlimited"),
    ("", "", "", "", ""),
    ("API cost per scan", "", "", "", ""),
    ("Monthly API cost (30d)", "", "", "", ""),
    ("Monthly revenue", "€0", "€29", "€79", "€199"),
    ("Gross profit per customer", "", "", "", ""),
    ("Gross margin", "", "", "", ""),
]

for i, (feat, *vals) in enumerate(features):
    row = r + 1 + i
    ws3.cell(row, 1, feat).font = BLACK_BOLD if feat in ("Monthly price", "API cost per scan", "Monthly API cost (30d)", "Monthly revenue", "Gross profit per customer", "Gross margin") else BLACK
    for j, v in enumerate(vals):
        if feat == "API cost per scan":
            # Link to API Costs sheet
            checks = [30, 75, 250, 500]
            ws3.cell(row, j + 2, f"={checks[j]}*'API Costs'!$C$16").font = BLACK
            ws3.cell(row, j + 2).number_format = EUR
        elif feat == "Monthly API cost (30d)":
            days = [14, 30, 30, 30]  # trial only 14 days
            checks = [30, 75, 250, 500]
            ws3.cell(row, j + 2, f"={checks[j]}*'API Costs'!$C$16*{days[j]}").font = BLACK
            ws3.cell(row, j + 2).number_format = EUR
        elif feat == "Gross profit per customer":
            rev = [0, 29, 79, 199]
            cost_row = row - 1  # monthly API cost row
            ws3.cell(row, j + 2, f"={rev[j]}-{get_column_letter(j+2)}{cost_row}").font = BLACK_BOLD
            ws3.cell(row, j + 2).number_format = EUR
        elif feat == "Gross margin":
            rev = [0, 29, 79, 199]
            profit_row = row - 1
            if rev[j] > 0:
                ws3.cell(row, j + 2, f"={get_column_letter(j+2)}{profit_row}/{rev[j]}").font = BLACK_BOLD
                ws3.cell(row, j + 2).number_format = PCT
                ws3.cell(row, j + 2).fill = GREEN_FILL
            else:
                ws3.cell(row, j + 2, "N/A").font = Font(color="999999", name="Arial")
        else:
            ws3.cell(row, j + 2, v).font = BLACK
            ws3.cell(row, j + 2).alignment = Alignment(horizontal="center")

# Key insight box
r2 = r + len(features) + 3
ws3.cell(r2, 1, "KEY INSIGHT").font = SECTION
ws3.cell(r2 + 1, 1, "Because we use Claude to SIMULATE other models' responses (not calling OpenAI/Google/Perplexity directly),").font = BLACK
ws3.cell(r2 + 2, 1, "all API costs run through a single Anthropic bill. This simplifies ops but means cost per check is ~€0.02.").font = BLACK
ws3.cell(r2 + 3, 1, "").font = BLACK
ws3.cell(r2 + 4, 1, "COST REDUCTION STRATEGIES:").font = SECTION
ws3.cell(r2 + 5, 1, "1. Use Claude Haiku for the simulation step (10x cheaper, ~€0.002/check)").font = BLACK
ws3.cell(r2 + 6, 1, "2. Cache common prompt responses (same prompt rarely changes day-to-day)").font = BLACK
ws3.cell(r2 + 7, 1, "3. Use batch API for non-urgent daily scans (50% discount)").font = BLACK
ws3.cell(r2 + 8, 1, "4. Eventually integrate direct model APIs (OpenAI, Google) for more authentic results AND lower costs").font = BLACK

# Save
output = "/sessions/exciting-focused-ride/mnt/cmo-ie/CMO.ie_Financial_Model.xlsx"
wb.save(output)
print(f"Saved to {output}")
